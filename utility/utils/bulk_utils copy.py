# def reupload_excel(filepath, model, model_mapping):

import json
import traceback
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from openpyxl import load_workbook
import datetime
import pathlib
import uuid
from django.conf import settings
from inventory.models import Category, Variant, Products, Attributes
import string
import random
from django.db.models.signals import pre_save
from django.dispatch import receiver
from django.utils.text import slugify
from rest_framework.decorators import action
from rest_framework import viewsets, status
from rest_framework.response import Response
from coreapp.pagination import paginate

class FilterGivenDate(viewsets.ModelViewSet):
    @paginate
    @action(detail=False, methods=['get'])
    def filter(self, request, *args, **kwargs):
        start = request.GET.get("start")
        end = request.GET.get("end")
        try:
            start_month, start_day,  start_year = start.split("/")
        except:
            return Response({"data": [], "error": "Start date is empty or Please use correct format,month/date/year"}, status=status.HTTP_404_NOT_FOUND)
        try:
            end_month, end_day,  end_year = end.split("/")
        except:
            return Response({"data": [], "error": "End date is empty or Please use correct format,month/date/year"}, status=status.HTTP_404_NOT_FOUND)
        start_date = datetime.datetime(year=int(start_year), month=int(start_month), day=int(start_day),
                                        hour=0, minute=0, second=0)  # represents 00:00:00
        end_date = datetime.datetime(year=int(end_year), month=int(end_month), day=int(end_day),
                                        hour=23, minute=59, second=59)
        return self.queryset.model.objects.filter(created_at__range=[start_date, end_date])




def random_string_generator(size=10, chars=string.ascii_lowercase + string.digits):
    return ''.join(random.choice(chars) for _ in range(size))


def unique_slug_generator(instance, new_slug=None):
    import uuid
    if new_slug is not None:
        slug = new_slug
    else:
        try:
            slug = slugify(instance.name)
        except:
            try:
                slug = slugify(instance.number)
            except:
                slug = slugify(str(uuid.uuid4()))

    Klass = instance.__class__
    max_length = Klass._meta.get_field('slug').max_length
    slug = slug[:max_length]
    qs_exists = Klass.objects.filter(slug=slug).exists()

    if qs_exists:
        new_slug = "{slug}-{randstr}".format(
            slug=slug[:max_length-5], randstr=random_string_generator(size=4))

        return unique_slug_generator(instance, new_slug=new_slug)
    return slug




def BulkImportFunction(filepath):
    wb = load_workbook(filepath)
    sheet = wb["Sheet1"]
    max_column = sheet.max_column
    max_row = sheet.max_row
    heading = []
    result = []
    images = []
    for row in range(1, max_row+1):
        current_data = {}
        new_path = ''

        for column in range(1, max_column+1):
            value = sheet.cell(row, column).value
            if row == 1:
                heading.append(value)
            if column == 2 and row != 0:
                try:
                    image_loader = SheetImageLoader(sheet)
                    image = image_loader.get('B' + str(row+1))
                    dir_path = f"product/bulk/"
                    # dir_path = f"product/{now.year}/{now.month}/"
                    pathlib.Path(settings.MEDIA_ROOT + f"/{dir_path}").mkdir(parents=True, exist_ok=True)
                    new_path = dir_path + f"{uuid.uuid4()}.webp"

                    # Image upload here to the directory
                    if image:
                        image.save(settings.MEDIA_ROOT + f"/{new_path}", "WEBP", optimize=False, quality=100)
                    # print(new_path)
                    images.append(new_path)
                    # current_data['thumb'] = new_path

                except Exception as e:
                    print("error ", e)
            else:
                try:
                    current_data[heading[column-1]] = value
                    current_data["row"] = row
                    current_data["column"] = column
                except:
                    traceback.print_exc()
                # result.append(value)
        result.append(current_data)
    result.pop(0)  # removing header values
    index = 0
    # print("result", result)
    for item in result:
        print("item", item)

        try:
            image = images[index]
        except Exception as e:
            print(str(e))

        try:
            row = item['row']
            column = item['column']
            name = item['name']
            category = item['category']
            parent_category = item['parent_category']
            quantity = item['quantity']
            stock = item['stock']
            sku = item['sku']
            existing_product = Products.objects.filter(sku=sku).first()
            if existing_product:
                return {
                    "msg":  f"Row : {row} Product : {name}" + " A product with this SKU already exists.",
                    # "error": str(e)
                }
            
            name = item['name']
            price = item['price']
            discount = item['discount']
            description = item['description']
            attributes = item['attribute']
            variants = item['variants']
            attribute_values = item['attribute_values']  # not necessary
            row_column_text = f"Row : {row} Product : {name}"
            if variants == None:
                print("--variant is nOne--")
                return {
                    "msg": row_column_text + "There's a problem while creating Variant.Please add the values carefully or check if the variant section is empty or not",
                    # "error": str(e)
                }

            # todo: if not found createing category and parent category

            # getCategory, created = Category.objects.get_or_create(name=category)
            try:
                getCategory = Category.objects.filter(name=category).first()
                if getCategory is None:
                    getCategory = Category.objects.create(name=category)
            except Exception as e:
                return {
                    "msg": row_column_text + "There's a problem creating category.Please add the values carefully",
                    "error": str(e)
                }
            try:
                prod = Products.objects.create(name=name, quantity=quantity, stock=stock, sku=sku, category=getCategory,
                                               price=price, discount=discount, description=description, thumb=image)
            except Exception as e:
                return {
                    "msg": row_column_text + "There's a problem creating Product.Please add the values carefully",
                    "error": str(e)
                }

            if parent_category is not None:

                try:
                    getparentCategory = Category.objects.filter(name=parent_category).first()
                    if getparentCategory is None:
                        getparentCategory = Category.objects.create(name=parent_category)
                    getCategory.parent_name = getparentCategory
                    getCategory.save()
                except Exception as e:
                    return {
                        "msg": row_column_text + "There's a problem creating category.Please add the values carefully,Delete The created product,and other instances  if any and try again.",
                        "error": str(e)
                    }

                # give the same name as required in frontend
            # data = {"mainState":variants,"AttributesInputValue":attribute_values}
            data = {}
            data['mainState'] = json.loads(attribute_values)
            data['AttributesInputValue'] = json.loads(variants)
            print("variants", variants)
            print("attribute_values", attribute_values)
            print("data", data)

            # todo: adding variants json and save
            prod.variants_json = data
            prod.save()

            # todo: creating attributes
            attributesList = attributes.split(",")
            print("Attributes list", attributesList)
            for attribute in attributesList:
                try:
                    getAttribute = Attributes.objects.filter(name=attribute).first()
                    if getAttribute is None:
                        getAttribute = Attributes.objects.create(name=attribute)

                except Exception as e:
                    prod.delete()
                    return {
                        "msg": row_column_text + "There's a problem creating Attributes.Please add the values carefully,Delete The created product and Delete The created product,and other instances if any and try again.",
                        "error": str(e)
                    }

            # todo:save json to product

            # todo: converted to json and creating product variant

            if variants is not None:
                converted_variant = json.loads(variants)
                try:
                    for key in converted_variant:
                        print(key)
                        price = converted_variant[key]["variant_price"]
                        stock = converted_variant[key]["variant_stock"]
                        is_active = converted_variant[key]["variant_action"]
                        ProductVariants.objects.create(product_id=prod.id, variant=key,
                                                       price=price, stock=stock, is_active=is_active)
                except Exception as e:
                    prod.delete()
                    return {
                        "msg": row_column_text + "There's a problem creating Variant.Please add the values carefully,Delete The created product,and other instances if any and try again.",
                        "error": str(e)
                    }
            else:
                prod.delete()
                return {
                    "msg": row_column_text + "There's a problem while creating Variant.Please add the values carefully or check if the variant section is empty or not",
                    "error": str(e)
                }
        except Exception as e:
            print(e)
            return {
                "msg": "There's a problem in " + row_column_text,
                "error": str(e)
            }
        index += 1
    return 1


#!Variant format
# text = {
#     "XL-Green": {
#         "price": 200,
#         "stock": 12
#     },
#     "XL-Red": {
#         "price": 200,
#         "stock": 12
#     },
#     "XXL-Red": {
#         "price": 200,
#         "stock": 12
#     },
#     "XXL-Green": {
#         "price": 200,
#         "stock": 12
#     }
# }

# json Store format


def CustomerBulkImport(filepath):
    from inventory.models import Customer
    wb = load_workbook(filepath)
    sheet = wb["Sheet1"]
    max_column = sheet.max_column
    max_row = sheet.max_row
    print(max_column)
    print(max_row)
    heading = []
    result = []
    images = []
    for row in range(1, max_row+1):
        current_data = {}
        new_path = ''

        for column in range(1, max_column+1):
            value = sheet.cell(row, column).value
            # print(value,end="  ")
            if row == 1:
                heading.append(value)

            else:
                try:
                    current_data[heading[column-1]] = value
                except:
                    traceback.print_exc()
        result.append(current_data)
    result.pop(0)  # removing header values
    index = 0
    print(result)
    try:
        for item in result:
            try:
                name = item['name']
                try:
                    mobile = "+880"+str(int(item['mobile']))
                except:
                    mobile = "+880"+item['mobile']

                email = item['email']
                address = item['address']
                customer = Customer.objects.create(name=name, mobile=mobile,
                                                   email=email, status=True, address=address)
            except Exception as e:
                return {
                    "msg": "There's a problem creating customer.Please add the values carefully",
                    "error": str(e)
                }
    except Exception as e:
        return {
            "msg": "There's a problem creating customer.Please add the values carefully",
            "error": str(e)
        }

    return 1






def duplicate_sku():
    from django.db.models import Count
    duplicates = Products.objects.values('sku').annotate(count=Count('sku'))
    for item in duplicates.order_by('-count'):
         if item['count']!=1:
            sku = item['sku']
            count = item['count']
            products = Products.objects.filter(sku=sku)
            print(f"SKU '{sku}' has {count} products:")
            for index,product in enumerate(products):
                print(f"{index+1} - {product.name} : URL : https://inventory.kaaruj.cloud/inventory/edit-product/{product.slug})")
            print("----------------------------------------------------")